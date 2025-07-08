import os
import pandas as pd
from telegram import Update
from telegram.ext import ApplicationBuilder, ContextTypes, CommandHandler
import datetime

TOKEN = os.getenv("BOT_TOKEN", "your_token_here")  # 建議你改返環境變數
CHAT_ID = os.getenv("CHAT_ID", "214241911")        # 建議你改返環境變數
URL = os.getenv("WEBHOOK_URL", "https://your-domain.onrender.com/")  # 記得最後有 "/"

from openpyxl import Workbook

# 建立假報表，方便測試傳送
def create_report(path='output/3t_report.xlsx'):
    wb = Workbook()
    ws = wb.active
    ws.title = "3T報表"
    ws['A1'] = '馬匹'
    ws['B1'] = '值搏率'
    ws.append(['喜蓮勇感', 1.8])
    ws.append(['紅衣醒神', 2.3])
    wb.save(path)

# /get3t 指令
async def get3t_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    file_path = 'output/3t_report.xlsx'
    if not os.path.exists(file_path):
        create_report(file_path)
    await context.bot.send_document(
        chat_id=update.effective_chat.id,
        document=open(file_path, 'rb'),
        filename="3t_report.xlsx",
        caption="📊 今日3T報表送上！"
    )

# 建立 bot app
app = ApplicationBuilder().token(TOKEN).build()

# 指令處理器
app.add_handler(CommandHandler("get3t", get3t_handler))

# webhook 模式（Render 用）
app.run_webhook(
    listen="0.0.0.0",
    port=10000,
    webhook_url=URL + TOKEN,
    secret_token=TOKEN
)
